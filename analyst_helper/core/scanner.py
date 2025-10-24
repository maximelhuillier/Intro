"""
Module de scan d'arborescence de dossiers
Permet de parcourir récursivement les dossiers et collecter les métadonnées
"""

import os
import hashlib
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Optional
from dataclasses import dataclass, asdict
import json


@dataclass
class FileInfo:
    """Information sur un fichier"""
    name: str
    path: str
    relative_path: str
    extension: str
    size_bytes: int
    size_kb: float
    size_mb: float
    created_date: str
    modified_date: str
    file_type: str
    depth: int
    parent_folder: str
    is_email: bool = False
    email_source: Optional[str] = None

    def to_dict(self) -> Dict:
        """Convertit en dictionnaire"""
        return asdict(self)


class FolderScanner:
    """Scanner d'arborescence de dossiers"""

    # Extensions par catégorie
    TECHNICAL_EXTENSIONS = {'.pdf', '.dwg', '.dxf', '.doc', '.docx', '.xls', '.xlsx',
                           '.ppt', '.pptx', '.odt', '.ods', '.odp'}
    EMAIL_EXTENSIONS = {'.msg', '.eml'}
    ARCHIVE_EXTENSIONS = {'.zip', '.rar', '.7z', '.tar', '.gz'}
    IMAGE_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.svg'}

    def __init__(self, root_path: str, max_depth: int = 50):
        """
        Initialise le scanner

        Args:
            root_path: Chemin racine à scanner
            max_depth: Profondeur maximale de récursion
        """
        self.root_path = Path(root_path).resolve()
        self.max_depth = max_depth
        self.files: List[FileInfo] = []
        self.stats = {
            'total_files': 0,
            'total_size_mb': 0,
            'total_emails': 0,
            'total_folders': 0,
            'by_type': {},
            'by_extension': {}
        }

    def scan(self, exclude_folders: Optional[List[str]] = None) -> List[FileInfo]:
        """
        Scanne le dossier racine

        Args:
            exclude_folders: Liste des dossiers à exclure

        Returns:
            Liste des fichiers trouvés
        """
        if exclude_folders is None:
            exclude_folders = []

        print(f"🔍 Scan de: {self.root_path}")
        self.files = []
        self._scan_directory(self.root_path, 0, exclude_folders)
        self._compute_stats()

        print(f"✅ Scan terminé: {self.stats['total_files']} fichiers trouvés")
        return self.files

    def _scan_directory(self, directory: Path, depth: int, exclude_folders: List[str]):
        """
        Scanne un dossier récursivement

        Args:
            directory: Dossier à scanner
            depth: Profondeur actuelle
            exclude_folders: Dossiers à exclure
        """
        if depth > self.max_depth:
            print(f"⚠️  Profondeur maximale atteinte pour: {directory}")
            return

        try:
            for entry in os.scandir(directory):
                # Ignorer les dossiers exclus
                if entry.name in exclude_folders:
                    continue

                if entry.is_file():
                    file_info = self._create_file_info(entry, depth)
                    if file_info:
                        self.files.append(file_info)

                elif entry.is_dir():
                    self.stats['total_folders'] += 1
                    self._scan_directory(Path(entry.path), depth + 1, exclude_folders)

        except PermissionError:
            print(f"⚠️  Accès refusé: {directory}")
        except Exception as e:
            print(f"❌ Erreur lors du scan de {directory}: {e}")

    def _create_file_info(self, entry: os.DirEntry, depth: int) -> Optional[FileInfo]:
        """
        Crée un objet FileInfo à partir d'une entrée

        Args:
            entry: Entrée du système de fichiers
            depth: Profondeur dans l'arborescence

        Returns:
            FileInfo ou None en cas d'erreur
        """
        try:
            stat = entry.stat()
            file_path = Path(entry.path)
            extension = file_path.suffix.lower()

            # Déterminer le type de fichier
            file_type = self._classify_file(extension)
            is_email = extension in self.EMAIL_EXTENSIONS

            return FileInfo(
                name=entry.name,
                path=str(file_path),
                relative_path=str(file_path.relative_to(self.root_path)),
                extension=extension,
                size_bytes=stat.st_size,
                size_kb=round(stat.st_size / 1024, 2),
                size_mb=round(stat.st_size / (1024 * 1024), 2),
                created_date=datetime.fromtimestamp(stat.st_ctime).strftime('%Y-%m-%d %H:%M:%S'),
                modified_date=datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S'),
                file_type=file_type,
                depth=depth,
                parent_folder=file_path.parent.name,
                is_email=is_email
            )

        except Exception as e:
            print(f"⚠️  Erreur lecture fichier {entry.name}: {e}")
            return None

    def _classify_file(self, extension: str) -> str:
        """
        Classifie un fichier selon son extension

        Args:
            extension: Extension du fichier

        Returns:
            Type de fichier
        """
        if extension in self.TECHNICAL_EXTENSIONS:
            return "Dossier technique"
        elif extension in self.EMAIL_EXTENSIONS:
            return "Correspondance"
        elif extension in self.ARCHIVE_EXTENSIONS:
            return "Archive"
        elif extension in self.IMAGE_EXTENSIONS:
            return "Image"
        else:
            return "Autres fichiers"

    def _compute_stats(self):
        """Calcule les statistiques sur les fichiers scannés"""
        self.stats['total_files'] = len(self.files)
        self.stats['total_size_mb'] = sum(f.size_mb for f in self.files)
        self.stats['total_emails'] = sum(1 for f in self.files if f.is_email)

        # Stats par type
        for file in self.files:
            self.stats['by_type'][file.file_type] = \
                self.stats['by_type'].get(file.file_type, 0) + 1
            self.stats['by_extension'][file.extension] = \
                self.stats['by_extension'].get(file.extension, 0) + 1

    def export_to_json(self, output_path: str):
        """
        Exporte les résultats en JSON

        Args:
            output_path: Chemin du fichier de sortie
        """
        data = {
            'root_path': str(self.root_path),
            'scan_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'stats': self.stats,
            'files': [f.to_dict() for f in self.files]
        }

        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)

        print(f"📄 Export JSON: {output_path}")

    def export_to_csv(self, output_path: str):
        """
        Exporte les résultats en CSV

        Args:
            output_path: Chemin du fichier de sortie
        """
        import csv

        if not self.files:
            print("⚠️  Aucun fichier à exporter")
            return

        with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.DictWriter(f, fieldnames=self.files[0].to_dict().keys())
            writer.writeheader()
            for file in self.files:
                writer.writerow(file.to_dict())

        print(f"📊 Export CSV: {output_path}")

    def export_to_excel(self, output_path: str):
        """
        Exporte les résultats en Excel (.xlsx)

        Args:
            output_path: Chemin du fichier de sortie
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            print("⚠️  Module 'openpyxl' non installé")
            print("   Installation : pip install openpyxl")
            return

        if not self.files:
            print("⚠️  Aucun fichier à exporter")
            return

        # Créer le workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Liste des fichiers"

        # En-têtes
        headers = [
            "Type", "Nom", "Extension", "Taille (KB)", "Taille (MB)",
            "Date création", "Date modification", "Dossier parent", "Chemin complet"
        ]

        # Style des en-têtes
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Écrire les en-têtes
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Écrire les données
        for row, file in enumerate(self.files, 2):
            ws.cell(row=row, column=1, value=file.file_type)
            ws.cell(row=row, column=2, value=file.name)
            ws.cell(row=row, column=3, value=file.extension)
            ws.cell(row=row, column=4, value=file.size_kb)
            ws.cell(row=row, column=5, value=file.size_mb)
            ws.cell(row=row, column=6, value=file.created_date)
            ws.cell(row=row, column=7, value=file.modified_date)
            ws.cell(row=row, column=8, value=file.parent_folder)
            ws.cell(row=row, column=9, value=file.path)

            # Bordures pour toutes les cellules
            for col in range(1, 10):
                ws.cell(row=row, column=col).border = border

        # Ajuster la largeur des colonnes
        column_widths = {
            'A': 20,  # Type
            'B': 40,  # Nom
            'C': 12,  # Extension
            'D': 12,  # Taille KB
            'E': 12,  # Taille MB
            'F': 20,  # Date création
            'G': 20,  # Date modification
            'H': 25,  # Dossier parent
            'I': 60   # Chemin complet
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Figer la première ligne
        ws.freeze_panes = 'A2'

        # Ajouter une feuille de statistiques
        ws_stats = wb.create_sheet("Statistiques")

        stats_data = [
            ["📊 Statistiques d'analyse", ""],
            ["", ""],
            ["Dossier analysé:", str(self.root_path)],
            ["Date d'analyse:", datetime.now().strftime('%d/%m/%Y %H:%M:%S')],
            ["", ""],
            ["Nombre total de fichiers:", self.stats['total_files']],
            ["Taille totale (MB):", f"{self.stats['total_size_mb']:.2f}"],
            ["Nombre d'emails:", self.stats['total_emails']],
            ["Nombre de dossiers:", self.stats['total_folders']],
            ["", ""],
            ["Répartition par type:", ""],
        ]

        for file_type, count in self.stats['by_type'].items():
            stats_data.append([f"  {file_type}:", count])

        stats_data.append(["", ""])
        stats_data.append(["Top 10 extensions:", ""])

        top_extensions = sorted(
            self.stats['by_extension'].items(),
            key=lambda x: x[1],
            reverse=True
        )[:10]

        for ext, count in top_extensions:
            stats_data.append([f"  {ext}:", count])

        # Écrire les stats
        for row_idx, (label, value) in enumerate(stats_data, 1):
            ws_stats.cell(row=row_idx, column=1, value=label)
            ws_stats.cell(row=row_idx, column=2, value=value)

            # Style du titre
            if row_idx == 1:
                ws_stats.cell(row=row_idx, column=1).font = Font(bold=True, size=14, color="4472C4")

        ws_stats.column_dimensions['A'].width = 30
        ws_stats.column_dimensions['B'].width = 20

        # Sauvegarder
        wb.save(output_path)
        print(f"📗 Export Excel: {output_path}")

    def get_tree_structure(self, max_depth: int = 3) -> str:
        """
        Génère une représentation textuelle de l'arborescence

        Args:
            max_depth: Profondeur maximale à afficher

        Returns:
            Arborescence sous forme de texte
        """
        tree = {}

        for file in self.files:
            if file.depth > max_depth:
                continue

            parts = Path(file.relative_path).parts
            current = tree

            for part in parts[:-1]:
                if part not in current:
                    current[part] = {}
                current = current[part]

            if '_files' not in current:
                current['_files'] = []
            current['_files'].append(file.name)

        return self._format_tree(tree)

    def _format_tree(self, tree: dict, prefix: str = "", is_last: bool = True) -> str:
        """
        Formate l'arborescence pour l'affichage

        Args:
            tree: Dictionnaire représentant l'arborescence
            prefix: Préfixe pour l'indentation
            is_last: Si c'est le dernier élément

        Returns:
            Arborescence formatée
        """
        result = []
        items = [(k, v) for k, v in tree.items() if k != '_files']
        files = tree.get('_files', [])

        for i, (name, subtree) in enumerate(items):
            is_last_item = (i == len(items) - 1) and not files
            connector = "└── " if is_last_item else "├── "
            result.append(f"{prefix}{connector}📁 {name}")

            extension = "    " if is_last_item else "│   "
            result.append(self._format_tree(subtree, prefix + extension, is_last_item))

        for i, filename in enumerate(files):
            is_last_file = i == len(files) - 1
            connector = "└── " if is_last_file else "├── "
            result.append(f"{prefix}{connector}📄 {filename}")

        return "\n".join(result)
