"""
Export Excel avancé avec traçabilité complète
Liste tous les fichiers et PJ avec leur origine et destination
"""

from pathlib import Path
from datetime import datetime
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .workflow_manager import ProcessedFile


class ExcelExporter:
    """Exporteur Excel avancé"""

    def __init__(self, output_path: str):
        """
        Initialise l'exporteur

        Args:
            output_path: Chemin du fichier Excel de sortie
        """
        self.output_path = output_path

    def export(self, processed_files: List[ProcessedFile], stats: dict):
        """
        Exporte vers Excel

        Args:
            processed_files: Liste des fichiers traités
            stats: Statistiques du traitement
        """
        wb = Workbook()

        # Feuille 1 : Liste complète des fichiers
        self._create_files_sheet(wb, processed_files)

        # Feuille 2 : Statistiques
        self._create_stats_sheet(wb, stats, processed_files)

        # Feuille 3 : Pièces jointes uniquement
        self._create_attachments_sheet(wb, processed_files)

        # Supprimer la feuille par défaut
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # Sauvegarder
        wb.save(self.output_path)
        print(f"[INFO] Export Excel : {self.output_path}")

    def _create_files_sheet(self, wb: Workbook, files: List[ProcessedFile]):
        """Crée la feuille avec tous les fichiers"""
        ws = wb.active
        ws.title = "Tous les fichiers"

        # En-têtes
        headers = [
            "Type",
            "Nom du fichier",
            "Extension",
            "Taille (KB)",
            "Chemin d'origine",
            "Chemin de classement",
            "Est une PJ ?",
            "Email source",
            "Date de traitement"
        ]

        # Styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        attachment_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Écrire les en-têtes
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

        # Écrire les données
        for row, file in enumerate(files, 2):
            path_obj = Path(file.classified_path)
            extension = path_obj.suffix

            ws.cell(row=row, column=1, value=file.file_type)
            ws.cell(row=row, column=2, value=path_obj.name)
            ws.cell(row=row, column=3, value=extension)
            ws.cell(row=row, column=4, value=file.size_kb)
            ws.cell(row=row, column=5, value=file.original_path)
            ws.cell(row=row, column=6, value=file.classified_path)
            ws.cell(row=row, column=7, value="Oui" if file.is_attachment else "Non")
            ws.cell(row=row, column=8, value=file.source_email or "")
            ws.cell(row=row, column=9, value=file.processed_date)

            # Colorer les PJ
            if file.is_attachment:
                for col in range(1, 10):
                    ws.cell(row=row, column=col).fill = attachment_fill

            # Bordures
            for col in range(1, 10):
                ws.cell(row=row, column=col).border = border

        # Largeurs de colonnes
        column_widths = {
            'A': 20,  # Type
            'B': 40,  # Nom
            'C': 12,  # Extension
            'D': 12,  # Taille
            'E': 60,  # Origine
            'F': 60,  # Classement
            'G': 12,  # PJ?
            'H': 40,  # Email source
            'I': 20   # Date
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Figer la première ligne
        ws.freeze_panes = 'A2'

        # Filtres automatiques
        ws.auto_filter.ref = f"A1:I{len(files) + 1}"

    def _create_stats_sheet(self, wb: Workbook, stats: dict, files: List[ProcessedFile]):
        """Crée la feuille de statistiques"""
        ws = wb.create_sheet("Statistiques")

        stats_data = [
            ["📊 Statistiques de traitement", ""],
            ["", ""],
            ["Date de traitement:", datetime.now().strftime('%d/%m/%Y %H:%M:%S')],
            ["", ""],
            ["TOTAUX", ""],
            ["Fichiers traités:", stats.get('total_files', 0)],
            ["Emails traités:", stats.get('total_emails', 0)],
            ["Emails renommés:", stats.get('emails_renamed', 0)],
            ["Pièces jointes extraites:", stats.get('total_attachments', 0)],
            ["", ""],
            ["RÉPARTITION PAR CATÉGORIE", ""],
        ]

        for category, count in stats.get('by_category', {}).items():
            stats_data.append([f"  {category}:", count])

        stats_data.extend([
            ["", ""],
            ["TAILLE TOTALE", ""],
            [f"  Total (KB):", sum(f.size_kb for f in files)],
            [f"  Total (MB):", round(sum(f.size_kb for f in files) / 1024, 2)],
            ["", ""],
            ["PIÈCES JOINTES", ""],
            [f"  Nombre total:", sum(1 for f in files if f.is_attachment)],
            [f"  Taille (MB):", round(sum(f.size_kb for f in files if f.is_attachment) / 1024, 2)],
        ])

        # Écrire les stats
        for row_idx, (label, value) in enumerate(stats_data, 1):
            ws.cell(row=row_idx, column=1, value=label)
            ws.cell(row=row_idx, column=2, value=value)

            # Styles
            if row_idx == 1:
                ws.cell(row=row_idx, column=1).font = Font(bold=True, size=14, color="4472C4")
            elif label and label.isupper():
                ws.cell(row=row_idx, column=1).font = Font(bold=True, size=12)

        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20

    def _create_attachments_sheet(self, wb: Workbook, files: List[ProcessedFile]):
        """Crée une feuille dédiée aux pièces jointes"""
        ws = wb.create_sheet("Pièces jointes")

        # Filtrer les PJ
        attachments = [f for f in files if f.is_attachment]

        # En-têtes
        headers = [
            "Nom de la PJ",
            "Type",
            "Taille (KB)",
            "Email source",
            "Chemin de classement",
            "Date d'extraction"
        ]

        # Styles
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Écrire les en-têtes
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Écrire les données
        for row, att in enumerate(attachments, 2):
            ws.cell(row=row, column=1, value=Path(att.classified_path).name)
            ws.cell(row=row, column=2, value=att.file_type)
            ws.cell(row=row, column=3, value=att.size_kb)
            ws.cell(row=row, column=4, value=att.source_email or "")
            ws.cell(row=row, column=5, value=att.classified_path)
            ws.cell(row=row, column=6, value=att.processed_date)

            # Bordures
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = border

        # Largeurs
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 60
        ws.column_dimensions['F'].width = 20

        # Figer
        ws.freeze_panes = 'A2'

        # Filtres
        if attachments:
            ws.auto_filter.ref = f"A1:F{len(attachments) + 1}"
